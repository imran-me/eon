#!/usr/bin/env python3
"""
EON analytics service — the numeric/ML side of EON, in Python.

Called by the PHP server (lib/Py.php) or by cron with the dataset JSON:

    python3 eon.py forecast   --dataset ../storage/data/demo-dataset.json [--company 2] [--months 3]
    python3 eon.py anomalies  --dataset ...   [--company 2]
    python3 eon.py evaluate   --dataset ...   [--employee 12 | --all]
    python3 eon.py report     --dataset ...   --kind receivables|payables|payroll|pnl|attendance --out ../storage/data/report.xlsx
    python3 eon.py health

Everything works with the standard library alone; numpy / pandas / openpyxl are
used when installed (pip install -r requirements.txt) for speed and .xlsx output.
Output is always one JSON object on stdout.
"""
from __future__ import annotations
import argparse, json, math, sys, os, statistics, csv
from datetime import date, datetime, timedelta
from collections import defaultdict

try:
    import numpy as np  # optional
except Exception:  # pragma: no cover
    np = None
try:
    import openpyxl  # optional, for .xlsx
    from openpyxl.styles import Font, PatternFill, Alignment
except Exception:  # pragma: no cover
    openpyxl = None

MONTHS = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']


# ---------------------------------------------------------------- helpers
def load_dataset(path: str | None) -> dict:
    raw = sys.stdin.read() if (path in (None, '-', '')) else open(path, encoding='utf-8').read()
    d = json.loads(raw)
    d.setdefault('meta', {})
    d['meta'].setdefault('today', date.today().isoformat())
    return d


def in_company(row: dict, company: int | None, shared_ok: bool = False) -> bool:
    if company is None:
        return True
    c = row.get('company_id')
    return (shared_ok and c is None) or (c is not None and int(c) == company)


def acct_type(code: str) -> str:
    return {'1': 'asset', '2': 'liability', '3': 'equity', '4': 'income'}.get(str(code)[:1], 'expense')


def month_key(s: str) -> str:
    return str(s)[:7]


def add_months(mk: str, n: int) -> str:
    y, m = int(mk[:4]), int(mk[5:7])
    m += n
    y += (m - 1) // 12
    m = (m - 1) % 12 + 1
    return f'{y:04d}-{m:02d}'


def bdtk(n: float) -> str:
    a, s = abs(n), '−' if n < 0 else ''
    if a >= 1e7:
        return f"{s}৳{a / 1e7:.{0 if a >= 1e8 else 1}f} Cr".replace('.0 ', ' ')
    if a >= 1e5:
        return f"{s}৳{a / 1e5:.{0 if a >= 1e6 else 1}f} L".replace('.0 ', ' ')
    if a >= 1e3:
        return f"{s}৳{a / 1e3:.1f}k".replace('.0k', 'k')
    return f"{s}৳{a:,.0f}"


def linear_fit(ys: list[float]) -> tuple[float, float]:
    """least-squares slope/intercept over x = 0..n-1 (numpy if available)"""
    n = len(ys)
    if n == 0:
        return 0.0, 0.0
    if n == 1:
        return 0.0, ys[0]
    if np is not None:
        a, b = np.polyfit(np.arange(n), np.array(ys, dtype=float), 1)
        return float(a), float(b)
    xs = list(range(n))
    mx, my = sum(xs) / n, sum(ys) / n
    sxx = sum((x - mx) ** 2 for x in xs) or 1.0
    a = sum((x - mx) * (y - my) for x, y in zip(xs, ys)) / sxx
    return a, my - a * mx


# ---------------------------------------------------------------- monthly series from the ledger
def monthly_pl(d: dict, company: int | None) -> dict[str, dict]:
    out: dict[str, dict] = defaultdict(lambda: {'income': 0.0, 'direct': 0.0, 'opex': 0.0, 'finance': 0.0})
    for je in d.get('journal_entries', []):
        if company is not None and int(je.get('company_id') or 0) != company:
            continue
        mk = month_key(je['date'])
        for it in je.get('items', []):
            code = str(it.get('account_code', ''))
            t = acct_type(code)
            if t == 'income':
                out[mk]['income'] += float(it.get('credit', 0)) - float(it.get('debit', 0))
            elif t == 'expense':
                v = float(it.get('debit', 0)) - float(it.get('credit', 0))
                if code[:1] == '5':
                    out[mk]['direct'] += v
                elif code[:1] == '8':
                    out[mk]['finance'] += v
                else:
                    out[mk]['opex'] += v
    for mk, r in out.items():
        r['net'] = r['income'] - r['direct'] - r['opex'] - r['finance']
    return dict(sorted(out.items()))


def cash_total(d: dict, company: int | None) -> float:
    return sum(float(b.get('balance', 0)) for b in d.get('banks', []) if in_company(b, company))


# ---------------------------------------------------------------- forecast
def cmd_forecast(d: dict, company: int | None, months: int = 3) -> dict:
    today = d['meta']['today']
    cur = month_key(today)
    pl = monthly_pl(d, company)
    closed = [(mk, r) for mk, r in pl.items() if mk < cur]
    if len(closed) < 2:
        return {'ok': False, 'error': 'need at least two closed months of journal history'}
    hist_income = [r['income'] for _, r in closed]
    hist_out = [r['direct'] + r['opex'] + r['finance'] for _, r in closed]
    hist_net = [r['net'] for _, r in closed]
    a_i, b_i = linear_fit(hist_income)
    a_o, b_o = linear_fit(hist_out)
    n = len(closed)
    # seasonality: ratio of each month to the trend, folded by calendar month (light-touch with short history)
    seas: dict[int, list[float]] = defaultdict(list)
    for i, (mk, r) in enumerate(closed):
        trend = a_i * i + b_i
        if trend > 0:
            seas[int(mk[5:7])].append(r['income'] / trend)
    sd_net = statistics.pstdev(hist_net) if n > 1 else 0.0
    # partial current month → run-rate
    day = int(today[8:10]); dim = (date(int(today[:4]), int(today[5:7]) % 12 + 1, 1) - timedelta(days=1)).day if int(today[5:7]) < 12 else 31
    cur_income = pl.get(cur, {}).get('income', 0.0) * dim / max(1, day)
    proj = []
    cash = cash_total(d, company)
    running = cash
    for k in range(0, months + 1):
        mk = add_months(cur, k)
        idx = n + k
        f_income = a_i * idx + b_i
        f_out = a_o * idx + b_o
        s = seas.get(int(mk[5:7]))
        if s:
            f_income *= statistics.fmean(s)
        if k == 0 and cur_income > 0:
            f_income = 0.5 * f_income + 0.5 * cur_income   # blend the run-rate in for the running month
        f_net = f_income - f_out
        running += f_net if k > 0 else (f_net - pl.get(cur, {}).get('net', 0.0))
        proj.append({'month': mk, 'income': round(f_income), 'outflow': round(f_out), 'net': round(f_net), 'net_low': round(f_net - 1.28 * sd_net), 'net_high': round(f_net + 1.28 * sd_net), 'cash_end': round(running), 'label': MONTHS[int(mk[5:7]) - 1]})
    avg_out = statistics.fmean(hist_out) if hist_out else 0.0
    burn = statistics.fmean(hist_net) < 0
    runway = (cash / -statistics.fmean(hist_net)) if burn else None
    growth = (a_i / statistics.fmean(hist_income) * 100) if hist_income and statistics.fmean(hist_income) else 0.0
    speak = (f"Revenue trend: {'+' if growth >= 0 else ''}{growth:.1f}% per month over {n} closed months. Next month projects income {bdtk(proj[1]['income'])} and net {bdtk(proj[1]['net'])} (80% band {bdtk(proj[1]['net_low'])} to {bdtk(proj[1]['net_high'])}); cash by end of {proj[-1]['label']} about {bdtk(proj[-1]['cash_end'])}." if len(proj) > 1 else '')
    if burn and runway:
        speak += f" At the current burn, cash covers {runway:.1f} months."
    return {'ok': True, 'company': company, 'history_months': n, 'history': [{'month': mk, **{k: round(v) for k, v in r.items()}} for mk, r in closed], 'trend_income_per_month': round(a_i), 'growth_pct_per_month': round(growth, 1), 'projection': proj, 'cash_now': round(cash), 'avg_monthly_outflow': round(avg_out), 'months_of_cover': round(cash / avg_out, 1) if avg_out else None, 'burning': burn, 'runway_months': round(runway, 1) if runway else None, 'method': 'least-squares trend × calendar-month seasonality, 80% band from historical net σ', 'speak': speak}


# ---------------------------------------------------------------- anomalies
def cmd_anomalies(d: dict, company: int | None) -> dict:
    today = d['meta']['today']; cur = month_key(today); day = int(today[8:10])
    rows = [e for e in d.get('expenses', []) if in_company(e, company) and e.get('approval_status') != 'rejected']
    by_key: dict[tuple, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    partial: dict[tuple, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for e in rows:
        k = (e.get('company_id'), e.get('category') or 'Uncategorised'); mk = month_key(e['expense_date']); amt = float(e.get('amount', 0))
        by_key[k][mk] += amt
        if int(str(e['expense_date'])[8:10]) <= day:
            partial[k][mk] += amt
    months = sorted({month_key(e['expense_date']) for e in rows})
    co_name = {int(c['id']): (c.get('short_name') or c['name']) for c in d.get('companies', [])}
    out = []

    def flag(k, month, v, prior, projected):
        if len(prior) < 2:
            return
        mean = statistics.fmean(prior); sd = statistics.pstdev(prior) or mean * 0.1
        z = (v - mean) / sd if sd else 0
        if v > mean * 1.5 and z > 2 and v > 5000:
            out.append({'company': co_name.get(int(k[0]) if k[0] is not None else -1, str(k[0])), 'category': k[1], 'month': month, 'amount': round(v), 'projected': round(projected), 'usual': round(mean), 'ratio': round(v / mean, 2), 'z': round(z, 1)})

    for k, series in by_key.items():
        idx = months.index(cur) if cur in months else -1
        if idx >= 2:
            v = partial[k].get(cur, 0.0); prior = [partial[k].get(m, 0.0) for m in months[:idx] if partial[k].get(m, 0.0) > 0]
            flag(k, cur, v, prior, series.get(cur, 0.0) * 30 / max(1, day))
        li = idx - 1 if idx >= 0 else len(months) - 1
        if li >= 2:
            m = months[li]; v = series.get(m, 0.0); prior = [series.get(x, 0.0) for x in months[:li] if series.get(x, 0.0) > 0]
            flag(k, m, v, prior, v)
    # duplicate charges: same amount + category + company within 3 days
    seen = defaultdict(list); dups = []
    for e in rows:
        key = (e.get('company_id'), e.get('category'), round(float(e.get('amount', 0))))
        for prev in seen[key]:
            if abs((datetime.fromisoformat(str(e['expense_date'])[:10]) - datetime.fromisoformat(str(prev['expense_date'])[:10])).days) <= 3 and prev['id'] != e['id']:
                dups.append({'company': co_name.get(int(e.get('company_id') or -1), ''), 'category': e.get('category'), 'amount': round(float(e['amount'])), 'dates': [str(prev['expense_date'])[:10], str(e['expense_date'])[:10]], 'titles': [prev.get('title'), e.get('title')]})
                break
        seen[key].append(e)
    out.sort(key=lambda a: -a['ratio'])
    speak = (f"{len(out)} spending anomal{'ies' if len(out) != 1 else 'y'}: " + '; '.join(f"{a['category']} at {a['company']} {a['ratio']}× usual ({bdtk(a['amount'])} vs {bdtk(a['usual'])})" for a in out[:3])) if out else 'No spending anomalies — every category is inside its normal band.'
    if dups:
        speak += f" {len(dups)} possible duplicate charge(s) found."
    return {'ok': True, 'company': company, 'anomalies': out[:20], 'possible_duplicates': dups[:20], 'method': 'like-for-like day-of-month comparison vs prior months; z > 2 and 1.5× mean; duplicate = same company+category+amount within 3 days', 'speak': speak}


# ---------------------------------------------------------------- employee evaluation model
def evaluate_one(d: dict, e: dict, days: int = 30) -> dict:
    today = d['meta']['today']; frm = (date.fromisoformat(today) - timedelta(days=days)).isoformat(); uid = int(e['id'])
    att = [a for a in d.get('attendances', []) if int(a['user_id']) == uid and frm <= str(a['date']) <= today and a.get('status') != 'holiday']
    present = [a for a in att if a.get('status') == 'present']
    att_pct = len(present) / len(att) if att else 1.0
    late_days = sum(1 for a in present if int(a.get('late_minutes') or 0) > 0)
    punct = 1 - late_days / len(present) if present else 1.0
    ot = sum(int(a.get('overtime_minutes') or 0) for a in present)
    tasks = [t for t in d.get('tasks', []) if uid in (t.get('assigned_to') or [])]
    recent = [t for t in tasks if str(t.get('due_date') or '') >= frm or t.get('status') != 'done']
    done = [t for t in recent if t.get('status') == 'done']
    on_time = [t for t in done if not t.get('completed_at') or str(t['completed_at'])[:10] <= str(t.get('due_date') or '9999')]
    overdue = [t for t in recent if t.get('status') != 'done' and str(t.get('due_date') or '9999') < today]
    delivery = (len(on_time) / len(done)) if done else (0.6 if recent else None)
    leads = [l for l in d.get('leads', []) if int(l.get('assigned_to') or 0) == uid]
    won = sum(1 for l in leads if l.get('status') == 'won'); lost = sum(1 for l in leads if l.get('status') == 'lost')
    sales_rate = won / (won + lost) if (won + lost) else None
    is_sales = 'sales' in str(e.get('designation', '')).lower() or 'sales' in str(e.get('department', '')).lower()
    perf = sales_rate if (is_sales and sales_rate is not None) else (delivery if delivery is not None else 0.7)
    volume = min(1.0, (len(done) + won) / 6)
    # weights: attendance 30, punctuality 20, performance 35, volume 15  (+ overtime bonus up to 3, overdue penalty up to 5)
    score = att_pct * 30 + punct * 20 + perf * 35 + volume * 15 + min(3, ot / 600 * 3) - min(5, len(overdue) * 1.5)
    score = int(round(max(0, min(100, score))))
    grade = 'A' if score >= 85 else 'B' if score >= 70 else 'C' if score >= 55 else 'D'
    risk = []
    if att_pct < 0.85: risk.append('attendance')
    if punct < 0.7: risk.append('punctuality')
    if len(overdue) >= 2: risk.append('delivery')
    return {'id': uid, 'name': e.get('name'), 'designation': e.get('designation'), 'department': e.get('department'), 'company_id': e.get('company_id'), 'score': score, 'grade': grade, 'attendance_pct': round(att_pct * 100), 'punctuality_pct': round(punct * 100), 'late_days': late_days, 'overtime_hours': round(ot / 60, 1), 'tasks_done': len(done), 'tasks_on_time': len(on_time), 'tasks_overdue': len(overdue), 'open_tasks': len(recent) - len(done), 'leads_won': won, 'leads_lost': lost, 'sales_rate_pct': round(sales_rate * 100) if sales_rate is not None else None, 'attrition_risk': 'high' if len(risk) >= 2 else ('watch' if risk else 'low'), 'risk_factors': risk}


def cmd_evaluate(d: dict, company: int | None, employee: int | None, all_: bool) -> dict:
    emps = [e for e in d.get('employees', []) if in_company(e, company) and (e.get('status', 'active') == 'active')]
    if employee is not None:
        e = next((x for x in emps if int(x['id']) == employee), None)
        if not e:
            return {'ok': False, 'error': f'employee {employee} not found'}
        return {'ok': True, 'evaluation': evaluate_one(d, e)}
    rows = sorted((evaluate_one(d, e) for e in emps if int(e['id']) != 1), key=lambda r: -r['score'])
    dept = defaultdict(list)
    for r in rows:
        dept[r['department']].append(r['score'])
    return {'ok': True, 'count': len(rows), 'top': rows[:10], 'bottom': list(reversed(rows[-10:])), 'by_department': sorted(({'department': k, 'avg_score': round(statistics.fmean(v)), 'heads': len(v)} for k, v in dept.items()), key=lambda x: -x['avg_score']), 'attrition_risk_high': [r for r in rows if r['attrition_risk'] == 'high'][:10], 'distribution': {g: sum(1 for r in rows if r['grade'] == g) for g in 'ABCD'}, 'speak': (f"{len(rows)} staff evaluated: {sum(1 for r in rows if r['grade']=='A')} A, {sum(1 for r in rows if r['grade']=='B')} B, {sum(1 for r in rows if r['grade']=='C')} C, {sum(1 for r in rows if r['grade']=='D')} D. Top: {rows[0]['name']} ({rows[0]['score']}). {len([r for r in rows if r['attrition_risk']=='high'])} at high attrition risk." if rows else 'no staff')}


# ---------------------------------------------------------------- reports (xlsx / csv)
def report_rows(d: dict, company: int | None, kind: str) -> tuple[list[str], list[list]]:
    today = d['meta']['today']
    if kind in ('receivables', 'payables'):
        t = 'receive' if kind == 'receivables' else 'pay'
        rows = []
        for p in d.get('payment_schedules', []):
            if not in_company(p, company) or p.get('type') != t or p.get('status') not in ('pending', 'overdue'):
                continue
            due = float(p['amount']) - float(p.get('paid_amount') or 0); days = (date.fromisoformat(today) - date.fromisoformat(str(p['scheduled_date'])[:10])).days
            rows.append([p.get('party_name'), p.get('source_label'), str(p['scheduled_date'])[:10], round(due), max(0, days), p.get('status'), p.get('priority')])
        rows.sort(key=lambda r: (-r[4], -r[3]))
        return ['Party', 'Reference', 'Due date', 'Amount due', 'Days overdue', 'Status', 'Priority'], rows
    if kind == 'payroll':
        keys = sorted({p.get('month_key') for p in d.get('payroll', []) if p.get('month_key')}); mk = keys[-1] if keys else ''
        names = {int(e['id']): e for e in d.get('employees', [])}
        rows = [[names.get(int(p['user_id']), {}).get('name'), names.get(int(p['user_id']), {}).get('department'), p['gross_salary'], p['total_deductions'], p['late_deduction'], p['absent_deduction'], p['loan_deduction'], p['overtime_salary'], p['net_salary'], p['status']] for p in d.get('payroll', []) if p.get('month_key') == mk and in_company(p, company)]
        return ['Employee', 'Department', 'Gross', 'Deductions', 'Late', 'Absence', 'Loan', 'Overtime', 'Net', 'Status'], rows
    if kind == 'pnl':
        pl = monthly_pl(d, company)
        return ['Month', 'Income', 'Direct cost', 'Opex', 'Finance', 'Net'], [[mk, round(r['income']), round(r['direct']), round(r['opex']), round(r['finance']), round(r['net'])] for mk, r in pl.items()]
    if kind == 'attendance':
        rows = [[a['date'], next((e['name'] for e in d.get('employees', []) if int(e['id']) == int(a['user_id'])), a['user_id']), a.get('status'), a.get('check_in'), a.get('check_out'), a.get('late_minutes'), a.get('overtime_minutes')] for a in d.get('attendances', []) if in_company(a, company) and str(a['date']) >= (date.fromisoformat(today) - timedelta(days=30)).isoformat()]
        return ['Date', 'Employee', 'Status', 'In', 'Out', 'Late min', 'OT min'], rows
    raise ValueError('unknown report kind')


def cmd_report(d: dict, company: int | None, kind: str, out: str) -> dict:
    header, rows = report_rows(d, company, kind)
    os.makedirs(os.path.dirname(os.path.abspath(out)) or '.', exist_ok=True)
    if out.lower().endswith('.xlsx') and openpyxl is not None:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = kind[:31]
        ws.append([f'EON — {kind} report', f"generated {datetime.now():%Y-%m-%d %H:%M}"]); ws['A1'].font = Font(bold=True, size=13)
        ws.append([]); ws.append(header)
        for c in ws[3]:
            c.font = Font(bold=True, color='FFFFFF'); c.fill = PatternFill('solid', fgColor='4F46E5'); c.alignment = Alignment(horizontal='center')
        for r in rows:
            ws.append(r)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(12, min(48, max(len(str(c.value or '')) for c in col) + 2))
        ws.freeze_panes = 'A4'
        wb.save(out)
        fmt = 'xlsx'
    else:
        if out.lower().endswith('.xlsx'):
            out = out[:-5] + '.csv'
        with open(out, 'w', newline='', encoding='utf-8-sig') as f:
            w = csv.writer(f); w.writerow(header); w.writerows(rows)
        fmt = 'csv'
    return {'ok': True, 'kind': kind, 'rows': len(rows), 'file': os.path.abspath(out), 'format': fmt, 'speak': f'{kind} report ready — {len(rows)} rows ({fmt}).'}


# ---------------------------------------------------------------- main
def main(argv=None) -> int:
    try:
        sys.stdout.reconfigure(encoding='utf-8')  # JSON out is always UTF-8 (Windows consoles default to cp1252)
    except Exception:
        pass
    ap = argparse.ArgumentParser(description='EON analytics service')
    ap.add_argument('command', choices=['forecast', 'anomalies', 'evaluate', 'report', 'health'])
    ap.add_argument('--dataset', default='-'); ap.add_argument('--company', type=int); ap.add_argument('--months', type=int, default=3)
    ap.add_argument('--employee', type=int); ap.add_argument('--all', action='store_true'); ap.add_argument('--kind', default='receivables'); ap.add_argument('--out', default='report.xlsx')
    a = ap.parse_args(argv)
    try:
        if a.command == 'health':
            res = {'ok': True, 'python': sys.version.split()[0], 'numpy': bool(np), 'openpyxl': bool(openpyxl)}
        else:
            d = load_dataset(a.dataset)
            res = {'forecast': lambda: cmd_forecast(d, a.company, a.months), 'anomalies': lambda: cmd_anomalies(d, a.company), 'evaluate': lambda: cmd_evaluate(d, a.company, a.employee, a.all), 'report': lambda: cmd_report(d, a.company, a.kind, a.out)}[a.command]()
    except Exception as ex:  # always answer with JSON
        res = {'ok': False, 'error': f'{type(ex).__name__}: {ex}'}
    sys.stdout.write(json.dumps(res, ensure_ascii=False))
    return 0 if res.get('ok') else 1


if __name__ == '__main__':
    sys.exit(main())
